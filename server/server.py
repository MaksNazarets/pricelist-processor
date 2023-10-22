from datetime import datetime
import json
import os
from flask import Flask
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from flask import request, send_file

from my_utils import get_ws_with_category_column

app = Flask(__name__)

app.secret_key = 'some_secret_key!'

CORS(app)


@app.route('/get-sheets', methods=['POST'])
def get_sheets():
    file = request.files['file']

    wb = load_workbook(file)

    return wb.sheetnames, 200


@app.route('/get-cols', methods=['POST'])
def get_cols():
    file = request.files['file']
    sheet_name = request.form.get('sheetName')
    table_start = request.form.get('tableStart')

    wb = load_workbook(file)
    ws = wb[sheet_name]
    header_row_number, header_col_start = coordinate_to_tuple(table_start)

    # Get the table header column range
    header_column_range = ws.iter_cols(
        min_col=header_col_start, max_col=ws.max_column, min_row=header_row_number, max_row=header_row_number)

    # Get all the cell values in the header row
    header_values = []
    for cell in header_column_range:
        if cell[0].value and str(cell[0].value).strip() != '':
            header_values.append(cell[0].value)
        else:
            break

    has_category_rows = False

    table_body_column_range = ws.iter_rows(
        min_col=header_col_start, max_col=ws.max_column, min_row=header_row_number+1, max_row=ws.max_row)

    for row in table_body_column_range:
        if sum(1 for cell in row if cell.value) == 1:
            has_category_rows = True
            break

    return {'columns': header_values,
            'hasCategoryRows': has_category_rows}, 200


@app.route('/format-pricelist', methods=['POST'])
def format_pricelist():
    file = request.files['file']
    item_list = json.loads(request.form.get('itemList'))
    new_columns = json.loads(request.form.get('newColumns'))
    selected_sheet = request.form.get('selectedSheetName')
    table_start_cell = request.form.get('tableStartCell')
    cols_renamed = json.loads(request.form.get('columnsRenamed'))
    has_category_rows = request.form.get('hasCategoryRows')
    category_col_name = request.form.get('categoryColumnName')

    if not file:
        return 400

    old_wb = load_workbook(file)
    old_ws = old_wb[selected_sheet]

    new_wb = Workbook()
    new_ws = new_wb.active

    new_ws.title = selected_sheet

    old_start_row, old_start_col = coordinate_to_tuple(table_start_cell)

    if has_category_rows:
        old_ws = get_ws_with_category_column(old_ws, old_start_col, old_start_row, len(
            cols_renamed.keys()), category_col_name)

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(fill_type="solid", fgColor="B7DDE8")

    new_cols_names = [col['name'] for col in new_columns]

    new_cols_items = []

    for index, item_name in enumerate(item_list):
        new_column_letter = get_column_letter(index + 1)
        cell = new_ws[new_column_letter + "1"]
        cell.value = item_name

        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

        if item_name in new_cols_names:
            new_cols_items.append((index, item_name))

        else:
            old_col_index = len(cols_renamed.keys()) + 1  \
                if item_name == category_col_name \
                else list(
                cols_renamed.values()).index(item_name)

            old_col_letter = get_column_letter(old_start_col + old_col_index - 1) \
                if item_name == category_col_name \
                else get_column_letter(old_start_col + old_col_index)

            new_ws.column_dimensions[get_column_letter(
                index + 1)].width = old_ws.column_dimensions[old_col_letter].width

            # Assuming data starts from the second row
            for row in range(2, old_ws.max_row + 1):
                old_cell_value = old_ws[old_col_letter +
                                        str(old_start_row + row - 1)].value
                if old_cell_value:
                    new_ws.cell(row=row, column=index +
                                1, value=old_cell_value)

    repeats = []

    for i, (index, item_name) in enumerate(new_cols_items):
        col_index = new_cols_names.index(item_name)
        fillingMethod = new_columns[col_index]['fillingMethod']['name']

        new_ws.column_dimensions[get_column_letter(index + 1)].width = 25

        if fillingMethod == 'staticValue':
            value = new_columns[col_index]['fillingMethod']['value']
            for row in range(2, old_ws.max_row + 1):
                new_ws.cell(row=row, column=index +
                            1, value=value)

                # TODO: make the cells to be the same number as filled rows

        elif fillingMethod == 'formula':
            formula_text = new_columns[col_index]['fillingMethod']['formulaText']
            unique = new_columns[col_index]['fillingMethod']['unique']

            if unique:
                values = []

            process_rows = True
            for row in range(2, old_ws.max_row + 1):
                cell_val = ''
                col_scope_opened = False
                col_letter_index = ''
                for c in formula_text:
                    if c == '⌯':
                        if col_scope_opened:
                            col_scope_opened = False

                            letter_coords = f'{col_letter_index}{row}'
                            _, col = coordinate_to_tuple(letter_coords)

                            if (col - 1, item_list[col - 1]) in new_cols_items:

                                new_cols_items.append((index, item_name))

                                process_rows = False
                                break

                            cell_val += f"{new_ws[letter_coords].value or ''}"
                            col_letter_index = ''
                        else:
                            col_scope_opened = True

                    elif col_scope_opened:
                        col_letter_index += c

                    else:
                        cell_val += f'{c}'

                if not process_rows:
                    break

                new_ws.cell(row=row, column=index + 1, value=cell_val)

                # TODO: Implement unique value handling
                  # Треба робити прохід по готовій таблиці в окремому циклі перед кінцевим return
                # if unique:
                #   if cell_val in values:
                #       first_entry_row = values.index(cell_val) + 1
                #       if first_entry_row not in repeats:
                #           repeats.append(first_entry_row)
                #       repeats.append(row)
                #       print(f"repeats: {repeats}")
                #   else:
                #       values.append(cell_val)
                # else:
                #     values.append(cell_val)

                ######
        else:
            continue

        new_cols_items[i] = False

    filename = f'test_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    try:
        new_wb.save(os.path.join('files', filename))
    except Exception as e:
        print(e)
        return 'file save error', 501

    rows_with_repeats = []

    for row in repeats:
        row_cells = []
        for cell in new_ws[row]:
            row_cells.append(cell.value)

        rows_with_repeats.append({'cells': row_cells, 'row_number': row})
    return {
        'filename': filename,
        'repetitions': rows_with_repeats or None}, 200


@app.route('/download-file', methods=['GET'])
def download_file():
    filename = request.args.get('filename')

    return send_file(os.path.join('files', filename), as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)
