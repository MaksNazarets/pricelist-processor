from datetime import datetime
import os
from werkzeug.utils import secure_filename
from openpyxl.worksheet.worksheet import Worksheet

def save_file(file) -> str:
    filename = secure_filename(file.filename)
    base_path = 'files'
    save_path = os.path.join(base_path, filename)

    if os.path.exists(save_path):
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        file_name, file_extension = os.path.splitext(filename)
        new_filename = f"{file_name}_{current_time}{file_extension}"
        save_path = os.path.join(base_path, new_filename)

    file.save(save_path)

    return save_path


def get_ws_with_category_column(worksheet: Worksheet, start_column, start_row, cols_number, category_col_name):
    ws = worksheet
    ws.cell(row=start_row, column=start_column +
            cols_number).value = category_col_name

    category_column_number = start_column + cols_number

    category_name = ''

    category_rows = []
    for row in range(start_row + 1, ws.max_row + 1):
        row_cells = ws[row]
        if sum(1 for cell in row_cells if cell.value) == 1:  # if a category row
            col_with_text = 1
            for i, cell in enumerate(row_cells):
                if cell.value:
                    col_with_text = i + 1
                    break

            category_name = ws.cell(row=row, column=col_with_text).value

            try:
                ws.unmerge_cells(start_row=row, end_row=row,
                                 start_column=start_column, end_column=start_column+cols_number-1)
            except Exception as e:
                print(e)

            category_rows.append(row)
        else:
            ws.cell(row=row, column=category_column_number).value = category_name

    # removing category rows
    for i, r in enumerate(category_rows):
        ws.delete_rows(r - i)
        print(f'row {r} removed')

    return ws
